"""
Excel Export Service
Generates .xlsx files with job listings
"""
import os
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional
from pathlib import Path
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Generates Excel exports for job listings.
    Supports daily exports per run and master exports per user.
    """
    
    # Column definitions
    COLUMNS = [
        {"key": "id", "header": "Job ID", "width": 15},
        {"key": "source_name", "header": "Source", "width": 15},
        {"key": "company", "header": "Company", "width": 20},
        {"key": "title", "header": "Title", "width": 30},
        {"key": "location", "header": "Location", "width": 20},
        {"key": "region", "header": "Region", "width": 10},
        {"key": "remote_type", "header": "Remote Type", "width": 12},
        {"key": "seniority", "header": "Seniority", "width": 12},
        {"key": "posted_at", "header": "Posted Date", "width": 15},
        {"key": "scraped_at", "header": "Scraped Date", "width": 15},
        {"key": "match_score", "header": "Match Score", "width": 12},
        {"key": "matched_skills", "header": "Matched Skills", "width": 30},
        {"key": "matched_keywords", "header": "Matched Keywords", "width": 25},
        {"key": "salary_min", "header": "Salary Min", "width": 12},
        {"key": "salary_max", "header": "Salary Max", "width": 12},
        {"key": "salary_currency", "header": "Currency", "width": 10},
        {"key": "job_url", "header": "URL", "width": 50},
        {"key": "description_snippet", "header": "Description", "width": 50},
        {"key": "apply_url", "header": "Apply URL", "width": 50},
        {"key": "status", "header": "Status", "width": 10},
        {"key": "notes", "header": "Notes", "width": 30},
    ]
    
    # Styling
    HEADER_FILL = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SCORE_EXCELLENT_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    SCORE_GOOD_FILL = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    SCORE_FAIR_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    
    def __init__(self, export_path: str = None):
        self.export_path = export_path or os.environ.get('EXPORT_PATH', '/app/backend/exports')
        Path(self.export_path).mkdir(parents=True, exist_ok=True)
    
    def generate_export(
        self,
        jobs: List[Dict[str, Any]],
        user_id: str,
        export_type: str = "daily",
        run_id: Optional[str] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Generate an Excel export for jobs.
        
        Args:
            jobs: List of job dictionaries
            user_id: User ID
            export_type: "daily" (per run), "master" (all jobs), "filtered"
            run_id: Optional run ID for daily exports
            filters: Optional filters applied (for metadata)
        
        Returns:
            Dict with export info including filepath, filename, etc.
        """
        from app.models.schemas import generate_id, utc_now_iso
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"
        
        # Set up headers
        self._setup_headers(ws)
        
        # Add data rows
        for row_idx, job in enumerate(jobs, start=2):
            self._add_job_row(ws, row_idx, job)
        
        # Add summary sheet
        self._add_summary_sheet(wb, jobs, filters)
        
        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        if export_type == "daily" and run_id:
            filename = f"jobs_{export_type}_{run_id[:8]}_{timestamp}.xlsx"
        else:
            filename = f"jobs_{export_type}_{user_id[:8]}_{timestamp}.xlsx"
        
        filepath = os.path.join(self.export_path, filename)
        
        # Save workbook
        wb.save(filepath)
        file_size = os.path.getsize(filepath)
        
        # Create export record
        export_record = {
            "id": generate_id(),
            "user_id": user_id,
            "export_type": export_type,
            "run_id": run_id,
            "filename": filename,
            "filepath": filepath,
            "file_size": file_size,
            "job_count": len(jobs),
            "filters_applied": filters or {},
            "status": "completed",
            "error_message": "",
            "created_at": utc_now_iso(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        }
        
        logger.info(f"Generated export: {filename} with {len(jobs)} jobs")
        
        return export_record
    
    def generate_to_bytes(self, jobs: List[Dict[str, Any]], filters: Optional[Dict[str, Any]] = None) -> BytesIO:
        """Generate Excel file and return as BytesIO (for streaming response)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"
        
        self._setup_headers(ws)
        
        for row_idx, job in enumerate(jobs, start=2):
            self._add_job_row(ws, row_idx, job)
        
        self._add_summary_sheet(wb, jobs, filters)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _setup_headers(self, ws):
        """Set up header row with styling."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _add_job_row(self, ws, row_idx: int, job: Dict[str, Any]):
        """Add a job as a row in the worksheet."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            key = col_def["key"]
            value = job.get(key, "")
            
            # Handle special fields
            if key in ["matched_skills", "matched_keywords"]:
                value = ", ".join(value) if isinstance(value, list) else value
            elif key in ["posted_at", "scraped_at"]:
                if value:
                    try:
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        value = dt.strftime("%Y-%m-%d %H:%M")
                    except:
                        pass
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=key in ["description_snippet", "matched_skills"])
            
            # Color code match score
            if key == "match_score":
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value >= 80:
                    cell.fill = self.SCORE_EXCELLENT_FILL
                elif value >= 60:
                    cell.fill = self.SCORE_GOOD_FILL
                elif value >= 40:
                    cell.fill = self.SCORE_FAIR_FILL
    
    def _add_summary_sheet(self, wb, jobs: List[Dict[str, Any]], filters: Optional[Dict[str, Any]]):
        """Add a summary sheet with statistics."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws["A1"] = "Job Export Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:C1")
        
        # Stats
        stats = [
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Jobs", len(jobs)),
            ("", ""),
            ("By Status", ""),
        ]
        
        # Count by status
        status_counts = {}
        for job in jobs:
            status = job.get("status", "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
        
        for status, count in status_counts.items():
            stats.append((f"  - {status.capitalize()}", count))
        
        # Score distribution
        stats.append(("", ""))
        stats.append(("Match Score Distribution", ""))
        
        score_ranges = [
            ("80-100% (Excellent)", lambda s: s >= 80),
            ("60-79% (Good)", lambda s: 60 <= s < 80),
            ("40-59% (Fair)", lambda s: 40 <= s < 60),
            ("< 40% (Low)", lambda s: s < 40),
        ]
        
        for label, check in score_ranges:
            count = sum(1 for j in jobs if check(j.get("match_score", 0)))
            stats.append((f"  - {label}", count))
        
        # Average score
        if jobs:
            avg_score = sum(j.get("match_score", 0) for j in jobs) / len(jobs)
            stats.append(("", ""))
            stats.append(("Average Match Score", f"{avg_score:.1f}%"))
        
        # Top companies
        stats.append(("", ""))
        stats.append(("Top Companies", ""))
        company_counts = {}
        for job in jobs:
            company = job.get("company", "Unknown")
            company_counts[company] = company_counts.get(company, 0) + 1
        
        for company, count in sorted(company_counts.items(), key=lambda x: -x[1])[:10]:
            stats.append((f"  - {company}", count))
        
        # Filters applied
        if filters:
            stats.append(("", ""))
            stats.append(("Filters Applied", ""))
            for key, value in filters.items():
                stats.append((f"  - {key}", str(value)))
        
        # Write stats
        for row_idx, (label, value) in enumerate(stats, start=3):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25


async def create_export(
    db,
    user_id: str,
    export_type: str = "filtered",
    run_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    min_score: Optional[float] = None,
    source_filter: Optional[str] = None
) -> Dict[str, Any]:
    """
    Create an export from database jobs.
    """
    service = ExcelExportService()
    
    # Build query
    query = {"user_id": user_id}
    filters = {}
    
    if run_id:
        query["run_id"] = run_id
        filters["run_id"] = run_id
    
    if status_filter and status_filter != "all":
        query["status"] = status_filter
        filters["status"] = status_filter
    
    if min_score and min_score > 0:
        query["match_score"] = {"$gte": min_score}
        filters["min_score"] = min_score
    
    if source_filter:
        query["source_id"] = source_filter
        filters["source"] = source_filter
    
    # Fetch jobs
    jobs = await db.jobs.find(query, {"_id": 0}).sort("match_score", -1).to_list(10000)
    
    if not jobs:
        return {"error": "No jobs found matching criteria", "job_count": 0}
    
    # Generate export
    export_record = service.generate_export(
        jobs=jobs,
        user_id=user_id,
        export_type=export_type,
        run_id=run_id,
        filters=filters
    )
    
    # Save export record to DB
    await db.exports.insert_one(export_record)
    export_record.pop("_id", None)
    
    return export_record
