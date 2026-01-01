from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, BackgroundTasks, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.responses import StreamingResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import json
from io import BytesIO
import asyncio

# PDF and Document processing
import PyPDF2
from docx import Document

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# AI Integration
from anthropic import Anthropic

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'job-finder-secret-key-2025')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Anthropic AI Configuration
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# Create the main app
app = FastAPI(title="Job Finder AI System")

# Create routers
api_router = APIRouter(prefix="/api")
auth_router = APIRouter(prefix="/api/auth")
jobs_router = APIRouter(prefix="/api/jobs")
resume_router = APIRouter(prefix="/api/resume")
schedule_router = APIRouter(prefix="/api/schedule")
credentials_router = APIRouter(prefix="/api/credentials")
export_router = APIRouter(prefix="/api/export")

security = HTTPBearer()

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class ResumeProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    raw_text: str = ""
    parsed_data: Dict[str, Any] = {}
    skills: List[str] = []
    experience_years: int = 0
    roles: List[str] = []
    industries: List[str] = []
    education: List[str] = []
    location_preference: List[str] = []
    work_authorization: str = ""
    salary_range: Dict[str, Any] = {}
    remote_preference: str = "any"
    updated_at: str = ""

class JobPreferences(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    preferred_roles: List[str] = []
    preferred_industries: List[str] = []
    excluded_companies: List[str] = []
    included_companies: List[str] = []
    preferred_locations: List[str] = []
    min_salary: int = 0
    max_salary: int = 0
    seniority_levels: List[str] = []
    tech_stack: List[str] = []
    remote_only: bool = False
    posted_within_days: int = 30

class JobListing(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    source: str
    company: str
    title: str
    location: str
    remote_status: str = "unknown"
    posted_date: str = ""
    link: str
    match_score: float = 0.0
    matched_skills: List[str] = []
    salary_info: str = ""
    notes: str = ""
    status: str = "new"
    discovered_at: str = ""

class ScheduleConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    enabled: bool = True
    schedule_time: str = "07:30"
    regions: List[str] = []
    sources: List[str] = []
    frequency: str = "daily"
    last_run: str = ""
    next_run: str = ""

class CredentialVault(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    portal_name: str
    username: str
    encrypted_password: str = ""
    notes: str = ""
    last_used: str = ""

class PreferencesUpdate(BaseModel):
    preferred_roles: Optional[List[str]] = None
    preferred_industries: Optional[List[str]] = None
    excluded_companies: Optional[List[str]] = None
    included_companies: Optional[List[str]] = None
    preferred_locations: Optional[List[str]] = None
    min_salary: Optional[int] = None
    max_salary: Optional[int] = None
    seniority_levels: Optional[List[str]] = None
    tech_stack: Optional[List[str]] = None
    remote_only: Optional[bool] = None
    posted_within_days: Optional[int] = None

class ScheduleUpdate(BaseModel):
    enabled: Optional[bool] = None
    schedule_time: Optional[str] = None
    regions: Optional[List[str]] = None
    sources: Optional[List[str]] = None
    frequency: Optional[str] = None

class CredentialCreate(BaseModel):
    portal_name: str
    username: str
    password: str
    notes: Optional[str] = ""

class JobStatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {"user_id": user_id, "exp": expiration}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== AI HELPERS ====================

async def parse_resume_with_ai(text: str) -> Dict[str, Any]:
    """Use Claude to parse resume and extract structured data"""
    if not EMERGENT_LLM_KEY:
        return {"error": "AI key not configured", "raw_text": text}
    
    try:
        client = Anthropic(api_key=EMERGENT_LLM_KEY)
        
        prompt = f"""Analyze this resume and extract structured information. Return a JSON object with these fields:
- skills: list of technical and soft skills
- experience_years: estimated years of experience (number)
- roles: list of job titles/roles the person has held or is suitable for
- industries: list of industries they have experience in
- education: list of degrees/certifications
- location_preference: any mentioned location preferences
- work_authorization: visa/work authorization status if mentioned
- salary_range: any salary expectations mentioned (as object with min/max)
- remote_preference: "remote", "hybrid", "onsite", or "any"
- summary: brief professional summary
- keywords: important keywords for job matching

Resume text:
{text}

Return ONLY valid JSON, no additional text."""

        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text
        # Try to extract JSON from response
        try:
            # Handle potential markdown code blocks
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]
            return json.loads(response_text.strip())
        except json.JSONDecodeError:
            return {"raw_text": text, "parse_error": "Could not parse AI response"}
    except Exception as e:
        logging.error(f"AI parsing error: {str(e)}")
        return {"raw_text": text, "error": str(e)}

async def calculate_job_match_score(job: Dict, resume: Dict, preferences: Dict) -> Dict[str, Any]:
    """Calculate match score between job and resume/preferences"""
    if not EMERGENT_LLM_KEY:
        return {"score": 50, "matched_skills": [], "reason": "AI not configured"}
    
    try:
        client = Anthropic(api_key=EMERGENT_LLM_KEY)
        
        prompt = f"""Analyze the match between this job and candidate profile. Return a JSON object with:
- score: match percentage 0-100
- matched_skills: list of matching skills
- missing_skills: list of required skills candidate lacks
- reasons: brief explanation of score

Job:
Title: {job.get('title', '')}
Company: {job.get('company', '')}
Description: {job.get('description', '')}
Requirements: {job.get('requirements', '')}

Candidate Profile:
Skills: {resume.get('skills', [])}
Experience: {resume.get('experience_years', 0)} years
Roles: {resume.get('roles', [])}
Industries: {resume.get('industries', [])}

Preferences:
Preferred Roles: {preferences.get('preferred_roles', [])}
Tech Stack: {preferences.get('tech_stack', [])}
Remote: {preferences.get('remote_only', False)}

Return ONLY valid JSON."""

        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text
        try:
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]
            return json.loads(response_text.strip())
        except json.JSONDecodeError:
            return {"score": 50, "matched_skills": [], "reason": "Could not parse AI response"}
    except Exception as e:
        logging.error(f"Match scoring error: {str(e)}")
        return {"score": 50, "matched_skills": [], "reason": str(e)}

# ==================== AUTH ROUTES ====================

@auth_router.post("/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "password": hash_password(user_data.password),
        "created_at": now
    }
    
    await db.users.insert_one(user_doc)
    
    # Create default preferences
    await db.preferences.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "preferred_roles": [],
        "preferred_industries": [],
        "excluded_companies": [],
        "included_companies": [],
        "preferred_locations": [],
        "min_salary": 0,
        "max_salary": 0,
        "seniority_levels": [],
        "tech_stack": [],
        "remote_only": False,
        "posted_within_days": 30
    })
    
    # Create default schedule
    await db.schedules.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "enabled": False,
        "schedule_time": "07:30",
        "regions": ["US", "EU", "UK"],
        "sources": ["linkedin", "indeed", "glassdoor"],
        "frequency": "daily",
        "last_run": "",
        "next_run": ""
    })
    
    token = create_token(user_id)
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user_id, email=user_data.email, name=user_data.name, created_at=now)
    )

@auth_router.post("/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"])
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            created_at=user["created_at"]
        )
    )

@auth_router.get("/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(**user)

# ==================== RESUME ROUTES ====================

@resume_router.post("/upload")
async def upload_resume(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload and parse resume (PDF, DOCX, or TXT)"""
    filename = file.filename.lower()
    content = await file.read()
    
    text = ""
    
    if filename.endswith('.pdf'):
        pdf_reader = PyPDF2.PdfReader(BytesIO(content))
        for page in pdf_reader.pages:
            text += page.extract_text() or ""
    elif filename.endswith('.docx'):
        doc = Document(BytesIO(content))
        text = "\n".join([para.text for para in doc.paragraphs])
    elif filename.endswith('.txt'):
        text = content.decode('utf-8')
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use PDF, DOCX, or TXT")
    
    if not text.strip():
        raise HTTPException(status_code=400, detail="Could not extract text from file")
    
    # Parse with AI
    parsed_data = await parse_resume_with_ai(text)
    
    now = datetime.now(timezone.utc).isoformat()
    resume_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "raw_text": text,
        "parsed_data": parsed_data,
        "skills": parsed_data.get("skills", []),
        "experience_years": parsed_data.get("experience_years", 0),
        "roles": parsed_data.get("roles", []),
        "industries": parsed_data.get("industries", []),
        "education": parsed_data.get("education", []),
        "location_preference": parsed_data.get("location_preference", []),
        "work_authorization": parsed_data.get("work_authorization", ""),
        "salary_range": parsed_data.get("salary_range", {}),
        "remote_preference": parsed_data.get("remote_preference", "any"),
        "updated_at": now
    }
    
    # Upsert resume
    await db.resumes.update_one(
        {"user_id": user["id"]},
        {"$set": resume_doc},
        upsert=True
    )
    
    return {"message": "Resume uploaded and parsed successfully", "profile": resume_doc}

@resume_router.get("/profile")
async def get_resume_profile(user: dict = Depends(get_current_user)):
    """Get user's parsed resume profile"""
    resume = await db.resumes.find_one({"user_id": user["id"]}, {"_id": 0})
    if not resume:
        return {"message": "No resume uploaded yet", "profile": None}
    return {"profile": resume}

@resume_router.put("/profile")
async def update_resume_profile(
    updates: Dict[str, Any],
    user: dict = Depends(get_current_user)
):
    """Manually update resume profile data"""
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.resumes.update_one(
        {"user_id": user["id"]},
        {"$set": updates}
    )
    resume = await db.resumes.find_one({"user_id": user["id"]}, {"_id": 0})
    return {"message": "Profile updated", "profile": resume}

# ==================== JOBS ROUTES ====================

@jobs_router.get("/")
async def get_jobs(
    status: Optional[str] = None,
    min_score: Optional[float] = None,
    source: Optional[str] = None,
    limit: int = Query(50, le=200),
    skip: int = 0,
    user: dict = Depends(get_current_user)
):
    """Get discovered jobs with filters"""
    query = {"user_id": user["id"]}
    if status:
        query["status"] = status
    if min_score:
        query["match_score"] = {"$gte": min_score}
    if source:
        query["source"] = source
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("match_score", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.jobs.count_documents(query)
    
    return {"jobs": jobs, "total": total, "limit": limit, "skip": skip}

@jobs_router.get("/stats")
async def get_job_stats(user: dict = Depends(get_current_user)):
    """Get job statistics"""
    user_id = user["id"]
    
    total = await db.jobs.count_documents({"user_id": user_id})
    new_jobs = await db.jobs.count_documents({"user_id": user_id, "status": "new"})
    saved = await db.jobs.count_documents({"user_id": user_id, "status": "saved"})
    applied = await db.jobs.count_documents({"user_id": user_id, "status": "applied"})
    
    # Average match score
    pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": None, "avg_score": {"$avg": "$match_score"}}}
    ]
    result = await db.jobs.aggregate(pipeline).to_list(1)
    avg_score = result[0]["avg_score"] if result else 0
    
    # Jobs by source
    source_pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$source", "count": {"$sum": 1}}}
    ]
    sources = await db.jobs.aggregate(source_pipeline).to_list(20)
    
    return {
        "total": total,
        "new": new_jobs,
        "saved": saved,
        "applied": applied,
        "average_match_score": round(avg_score, 1),
        "by_source": {s["_id"]: s["count"] for s in sources if s["_id"]}
    }

@jobs_router.put("/{job_id}/status")
async def update_job_status(
    job_id: str,
    update: JobStatusUpdate,
    user: dict = Depends(get_current_user)
):
    """Update job status (new, saved, applied, rejected)"""
    update_doc = {"status": update.status}
    if update.notes:
        update_doc["notes"] = update.notes
    
    result = await db.jobs.update_one(
        {"id": job_id, "user_id": user["id"]},
        {"$set": update_doc}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {"message": "Job status updated"}

@jobs_router.delete("/{job_id}")
async def delete_job(job_id: str, user: dict = Depends(get_current_user)):
    """Delete a job from list"""
    result = await db.jobs.delete_one({"id": job_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"message": "Job deleted"}

# Sample job discovery (simulated - in production would use real scrapers)
@jobs_router.post("/discover")
async def discover_jobs(
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user)
):
    """Trigger job discovery based on resume and preferences"""
    
    # Get user's resume and preferences
    resume = await db.resumes.find_one({"user_id": user["id"]}, {"_id": 0})
    preferences = await db.preferences.find_one({"user_id": user["id"]}, {"_id": 0})
    
    if not resume:
        raise HTTPException(status_code=400, detail="Please upload your resume first")
    
    # Generate sample jobs (in production, this would scrape real job sites)
    sample_companies = [
        ("Google", "Software Engineer", "Mountain View, CA", "https://careers.google.com"),
        ("Microsoft", "Senior Developer", "Seattle, WA", "https://careers.microsoft.com"),
        ("Amazon", "Backend Engineer", "New York, NY", "https://amazon.jobs"),
        ("Meta", "Full Stack Developer", "Menlo Park, CA", "https://metacareers.com"),
        ("Apple", "iOS Developer", "Cupertino, CA", "https://jobs.apple.com"),
        ("Netflix", "Platform Engineer", "Los Gatos, CA", "https://jobs.netflix.com"),
        ("Spotify", "Data Engineer", "Remote", "https://spotifyjobs.com"),
        ("Stripe", "Software Engineer", "San Francisco, CA", "https://stripe.com/jobs"),
        ("Airbnb", "Frontend Engineer", "San Francisco, CA", "https://careers.airbnb.com"),
        ("Uber", "Backend Developer", "San Francisco, CA", "https://uber.com/careers"),
    ]
    
    now = datetime.now(timezone.utc).isoformat()
    new_jobs = []
    
    for company, title, location, link in sample_companies:
        # Check if job already exists
        existing = await db.jobs.find_one({
            "user_id": user["id"],
            "company": company,
            "title": title
        })
        
        if not existing:
            # Calculate match score with AI
            job_info = {
                "title": title,
                "company": company,
                "description": f"Exciting opportunity at {company} for a {title}",
                "requirements": f"Experience with {', '.join(resume.get('skills', [])[:5])}"
            }
            
            match_result = await calculate_job_match_score(
                job_info,
                resume,
                preferences or {}
            )
            
            job_doc = {
                "id": str(uuid.uuid4()),
                "user_id": user["id"],
                "source": "MNC Career Portal",
                "company": company,
                "title": title,
                "location": location,
                "remote_status": "hybrid" if "Remote" not in location else "remote",
                "posted_date": now[:10],
                "link": link,
                "match_score": match_result.get("score", 50),
                "matched_skills": match_result.get("matched_skills", []),
                "salary_info": "$120,000 - $200,000",
                "notes": "",
                "status": "new",
                "discovered_at": now
            }
            
            await db.jobs.insert_one(job_doc)
            new_jobs.append(job_doc)
    
    return {
        "message": f"Discovered {len(new_jobs)} new jobs",
        "new_jobs_count": len(new_jobs),
        "jobs": new_jobs
    }

# ==================== PREFERENCES ROUTES ====================

@api_router.get("/preferences")
async def get_preferences(user: dict = Depends(get_current_user)):
    """Get user's job preferences"""
    prefs = await db.preferences.find_one({"user_id": user["id"]}, {"_id": 0})
    return {"preferences": prefs}

@api_router.put("/preferences")
async def update_preferences(
    updates: PreferencesUpdate,
    user: dict = Depends(get_current_user)
):
    """Update job preferences"""
    update_dict = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    await db.preferences.update_one(
        {"user_id": user["id"]},
        {"$set": update_dict}
    )
    
    prefs = await db.preferences.find_one({"user_id": user["id"]}, {"_id": 0})
    return {"message": "Preferences updated", "preferences": prefs}

# ==================== SCHEDULE ROUTES ====================

@schedule_router.get("/")
async def get_schedule(user: dict = Depends(get_current_user)):
    """Get user's job hunting schedule"""
    schedule = await db.schedules.find_one({"user_id": user["id"]}, {"_id": 0})
    return {"schedule": schedule}

@schedule_router.put("/")
async def update_schedule(
    updates: ScheduleUpdate,
    user: dict = Depends(get_current_user)
):
    """Update job hunting schedule"""
    update_dict = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    await db.schedules.update_one(
        {"user_id": user["id"]},
        {"$set": update_dict}
    )
    
    schedule = await db.schedules.find_one({"user_id": user["id"]}, {"_id": 0})
    return {"message": "Schedule updated", "schedule": schedule}

# ==================== CREDENTIALS ROUTES ====================

@credentials_router.get("/")
async def get_credentials(user: dict = Depends(get_current_user)):
    """Get stored credentials (passwords hidden)"""
    creds = await db.credentials.find(
        {"user_id": user["id"]},
        {"_id": 0, "encrypted_password": 0}
    ).to_list(50)
    return {"credentials": creds}

@credentials_router.post("/")
async def add_credential(
    cred: CredentialCreate,
    user: dict = Depends(get_current_user)
):
    """Add new portal credential"""
    now = datetime.now(timezone.utc).isoformat()
    
    cred_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "portal_name": cred.portal_name,
        "username": cred.username,
        "encrypted_password": hash_password(cred.password),  # Simple encryption
        "notes": cred.notes,
        "last_used": now
    }
    
    await db.credentials.insert_one(cred_doc)
    
    # Return without password
    del cred_doc["encrypted_password"]
    return {"message": "Credential added", "credential": cred_doc}

@credentials_router.delete("/{cred_id}")
async def delete_credential(cred_id: str, user: dict = Depends(get_current_user)):
    """Delete a stored credential"""
    result = await db.credentials.delete_one({"id": cred_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Credential not found")
    return {"message": "Credential deleted"}

# ==================== EXPORT ROUTES ====================

@export_router.get("/excel")
async def export_to_excel(
    status: Optional[str] = None,
    min_score: Optional[float] = None,
    user: dict = Depends(get_current_user)
):
    """Export jobs to Excel file"""
    query = {"user_id": user["id"]}
    if status:
        query["status"] = status
    if min_score:
        query["match_score"] = {"$gte": min_score}
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("match_score", -1).to_list(1000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Source", "Company", "Role", "Location", "Remote", "Posted", "Match Score", "Skills Matched", "Salary", "Status", "Link", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        ws.cell(row=row_idx, column=1, value=job.get("source", ""))
        ws.cell(row=row_idx, column=2, value=job.get("company", ""))
        ws.cell(row=row_idx, column=3, value=job.get("title", ""))
        ws.cell(row=row_idx, column=4, value=job.get("location", ""))
        ws.cell(row=row_idx, column=5, value=job.get("remote_status", ""))
        ws.cell(row=row_idx, column=6, value=job.get("posted_date", ""))
        ws.cell(row=row_idx, column=7, value=job.get("match_score", 0))
        ws.cell(row=row_idx, column=8, value=", ".join(job.get("matched_skills", [])))
        ws.cell(row=row_idx, column=9, value=job.get("salary_info", ""))
        ws.cell(row=row_idx, column=10, value=job.get("status", ""))
        ws.cell(row=row_idx, column=11, value=job.get("link", ""))
        ws.cell(row=row_idx, column=12, value=job.get("notes", ""))
        
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    column_widths = [15, 20, 25, 20, 10, 12, 12, 30, 20, 10, 40, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"job_listings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== AVAILABLE JOB SOURCES ====================

@api_router.get("/sources")
async def get_job_sources():
    """Get available job sources and regions"""
    return {
        "sources": [
            {"id": "linkedin", "name": "LinkedIn", "type": "aggregator", "regions": ["global"]},
            {"id": "indeed", "name": "Indeed", "type": "aggregator", "regions": ["global"]},
            {"id": "glassdoor", "name": "Glassdoor", "type": "aggregator", "regions": ["global"]},
            {"id": "google_careers", "name": "Google Careers", "type": "mnc", "regions": ["global"]},
            {"id": "microsoft_careers", "name": "Microsoft Careers", "type": "mnc", "regions": ["global"]},
            {"id": "amazon_jobs", "name": "Amazon Jobs", "type": "mnc", "regions": ["global"]},
            {"id": "meta_careers", "name": "Meta Careers", "type": "mnc", "regions": ["global"]},
            {"id": "apple_jobs", "name": "Apple Jobs", "type": "mnc", "regions": ["global"]},
            {"id": "naukri", "name": "Naukri", "type": "regional", "regions": ["india"]},
            {"id": "seek", "name": "Seek", "type": "regional", "regions": ["australia", "sea"]},
            {"id": "stepstone", "name": "StepStone", "type": "regional", "regions": ["eu"]},
            {"id": "reed", "name": "Reed", "type": "regional", "regions": ["uk"]},
            {"id": "bayt", "name": "Bayt", "type": "regional", "regions": ["middle_east"]},
        ],
        "regions": [
            {"id": "us", "name": "United States"},
            {"id": "eu", "name": "European Union"},
            {"id": "uk", "name": "United Kingdom"},
            {"id": "india", "name": "India"},
            {"id": "middle_east", "name": "Middle East"},
            {"id": "sea", "name": "Southeast Asia"},
            {"id": "australia", "name": "Australia"},
            {"id": "canada", "name": "Canada"},
        ]
    }

# Health check
@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include all routers
app.include_router(api_router)
app.include_router(auth_router)
app.include_router(jobs_router)
app.include_router(resume_router)
app.include_router(schedule_router)
app.include_router(credentials_router)
app.include_router(export_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
